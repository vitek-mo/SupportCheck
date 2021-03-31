from selenium import webdriver
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from bs4 import BeautifulSoup


SITE_INPUT_MAX = 20
DEFAULT_DIR_PATH = os.path.abspath(os.curdir)
DEFAULT_FILE_NAME = "input.xlsx"
FULL_DEFAULT_PATH = DEFAULT_DIR_PATH + '\\' + DEFAULT_FILE_NAME
INPUT_PATH_REQUEST_MSG = "Input file path or press enter to accept default path (" + FULL_DEFAULT_PATH + "): "


def get_path():
    response = input(INPUT_PATH_REQUEST_MSG)
    if response:
        return response
    return FULL_DEFAULT_PATH


def load_from_excel(input_file_path):
    wb = load_workbook(input_file_path)
    ws = wb.active
    sn_w_header = ()
    for row in ws.values:
        for value in row:
            sn_w_header += (value,)
    return sn_w_header[1:]

def save(save_file_path, data):
    wb_result = Workbook()
    ws = wb_result.active
    ws['A1'] = 'Serial Number'
    ws['B1'] = 'Product Number'
    ws['C1'] = 'SAID'
    ws['D1'] = 'Contract Status'
    ws['E1'] = 'Contract start date'
    ws['F1'] = 'Contract finish date'
    ws['G1'] = 'Warranty Status'
    ws['H1'] = 'Warranty start date'
    ws['I1'] = 'Warranty finish date'
    index = 1
    for row in data:
        index += 1
        ws['A' + str(index)] = row[0]
        ws['B' + str(index)] = row[1]
        ws['C' + str(index)] = row[2]
        ws['D' + str(index)] = row[3]
        ws['E' + str(index)] = row[4]
        ws['F' + str(index)] = row[5]
        ws['G' + str(index)] = row[6]
        ws['H' + str(index)] = row[7]
        ws['I' + str(index)] = row[8]
    wb_result.save(save_file_path)


def prepare_data(data):
    sn_iso = ()
    strings_count = 0
    for item in data:
        sn_iso += (item + "  KZ",)
    if len(sn_iso) % SITE_INPUT_MAX == 0:
        strings_count = len(sn_iso) // SITE_INPUT_MAX
    else:
        strings_count = len(sn_iso) // SITE_INPUT_MAX + 1

    tuples = ()
    for i in range(strings_count):
        if i == strings_count - 1:
            sub_tuple = sn_iso[i * SITE_INPUT_MAX:]
        else:
            sub_tuple = sn_iso[i * SITE_INPUT_MAX: (i + 1) * SITE_INPUT_MAX]
        tuples = tuples + (sub_tuple,)

    result = ()
    for chunk in tuples:
        string = ''
        for item in chunk:
            string += item + "\n"
        result += (string[:-1],)
    return result


def open_site(link):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(link)
    return driver


def get_single_chunk_result(driver, chunk):
    list_search_tab = driver.find_element_by_xpath('//*[@id="tabs"]/ul/li[3]/a')
    bulk_input_button_xpath = '/html/body/div[3]/div/div/div/div[3]/div/form/table/tbody/tr[5]/td/div/button[3]'
    show_bulk_input_button = driver.find_element_by_xpath(bulk_input_button_xpath)
    bulk_input_textfield = driver.find_element_by_xpath('//*[@id="inputcvs"]')
    load_button = driver.find_element_by_xpath('//*[@id="loadData"]')
    submit_button = driver.find_element_by_xpath('//*[@id="submit"]')
    driver.implicitly_wait(30)
    list_search_tab.click()
    show_bulk_input_button.click()
    bulk_input_textfield.click()
    bulk_input_textfield.send_keys(chunk)
    load_button.click()
    submit_button.click()
    time.sleep(60)
    result = driver.page_source
    driver.refresh()
    return result


def get_said(input_data):
    results = []
    for sn in input_data:
        sn_id_number = int(''.join(filter(str.isdigit, sn.attrs['id'])))
        sntag_id = "obligationTop_" + str(sn_id_number) + "-obligation0"
        if not(is_unit_found(sn)):
            results += ('not found',)
        elif not(has_contract(sn)):
            results += ('no contract',)
        elif sn.find(id=sntag_id):
            said = sn.find(id=sntag_id).table.tbody.tr.findAll('td')[1].get_text()
            results += (said,)
        else:
            results += ("-",)
    return results


def divide_html_by_sn(input_data):
    result = ()
    for i in range(20):
        if input_data.find(id="obligationTop_" + str(i)):
            result += (input_data.find(id="obligationTop_" + str(i)),)
    return result


def is_unit_found(input_data):
    strings = [text for text in input_data.div.table.tbody.stripped_strings]
    return not('301' in strings)


def get_sn_states_messages(input_data):
    hrefs = input_data.findAll('a', {"href": "#"})
    results = []
    for chunk in hrefs:
        if "SN:" in chunk.text:
            results += (chunk.text,)
    return results


def get_pns(input_data):
    results = []
    for sn in input_data:
        if is_unit_found(sn):
            results += (sn.findAll('td')[-6].get_text(),)
        else:
            results += ('not found',)
    return results


def has_contract(input_data):
    return len(input_data.findAll('h3')) >= 6


def get_sns(input_data):
    results = []
    for chunk in input_data:
        sn = chunk.split('|')[0][4:].strip()
        results += (sn,)
    return results


def get_contract_states(input_data):
    results = []
    for message in input_data:
        if len(message.split('|')) > 1:
            contract_part = message.split('|')[1]
            if contract_part.split(' - ')[0] == 'Contract' and contract_part.split(' - ')[1] == 'Active':
                results += ('Active',)
            else:
                results += ('Inactive',)
        else:
            results += ('Not found',)
    return results


def get_contract_start_date(input_data):
    results = []
    for message in input_data:
        if len(message.split('|'))>1:
            contract_part = message.split('|')[1]
            if contract_part.split(' - ')[0] == 'Contract':
                results += (contract_part.split(' - ')[2].split(' to ')[0],)
            else:
                results += ('-',)
        else:
            results += ('-',)
    return results


def get_contract_finish_date(input_data):
    results = []
    for message in input_data:
        if len(message.split('|'))>1:
            contract_part = message.split('|')[1]
            if contract_part.split(' - ')[0] == 'Contract':
                results += (contract_part.split(' - ')[2].split(' to ')[1],)
            else:
                results += ('-',)
        else:
            results += ('-',)
    return results


def get_warranty_states(input_data):
    results = []
    for message in input_data:
        if len(message.split('|')) > 1:
            if message.split('|')[1].split(' - ')[0] == 'Warranty':
                warranty_part = message.split('|')[1]
            else:
                warranty_part = message.split('|')[2]
            if warranty_part.split(' - ')[1] == 'Active':
                results += ('Active',)
            else:
                results += ('Inactive',)
        else:
            results += ('Not found',)
    return results


def get_warranty_start_date(input_data):
    results = []
    for message in input_data:
        if len(message.split('|')) > 1:
            if message.split('|')[1].split(' - ')[0] == 'Warranty':
                warranty_part = message.split('|')[1]
            else:
                warranty_part = message.split('|')[2]
            results += (warranty_part.split(' - ')[2].split(' to ')[0],)
        else:
            results += ('Not found',)
    return results


def get_warranty_finish_date(input_data):
    results = []
    for message in input_data:
        if len(message.split('|')) > 1:
            if message.split('|')[1].split(' - ')[0] == 'Warranty':
                warranty_part = message.split('|')[1]
            else:
                warranty_part = message.split('|')[2]
            results += (warranty_part.split(' - ')[2].split(' to ')[1],)
        else:
            results += ('Not found',)
    return results


def compose(input_data):
    result = ()
    indices = tuple(range(len(input_data[0])))
    for i in indices:
        result += ((input_data[0][i], input_data[1][i], input_data[2][i],
                   input_data[3][i], input_data[4][i], input_data[5][i],
                   input_data[6][i], input_data[7][i], input_data[8][i]),)
    return result


path = get_path()

# read SNs from excel
whole_sns = load_from_excel(path)
prepared_input = prepare_data(whole_sns)

# open site with webdriver
SITE_LINK = 'https://obligation-ui.corp.int.hpe.com/ui-obligation-1.1/main/ObligationViewer.jsp#'
browser = open_site(SITE_LINK)

composed = ()
step = 0

for chunk in prepared_input:
    step += 1
    print('Step ' + str(step) + ' of ' + str(len(prepared_input)))
    html = get_single_chunk_result(browser, chunk)
    soup = BeautifulSoup(html, 'html.parser')

    sns_data = divide_html_by_sn(soup)
    sn_states = get_sn_states_messages(soup)
    sns = get_sns(sn_states)
    saids = get_said(sns_data)
    pns = get_pns(sns_data)

    contract_states = get_contract_states(sn_states)
    contract_start = get_contract_start_date(sn_states)
    contract_finish = get_contract_finish_date(sn_states)

    warranty_states = get_warranty_states(sn_states)
    warranty_start = get_warranty_start_date(sn_states)
    warranty_finish = get_warranty_finish_date(sn_states)

    composed += compose((sns, pns, saids, contract_states,
                        contract_start, contract_finish, warranty_states,
                        warranty_start, warranty_finish))


def print_composed(input_data):
    print('SN\t\t\tPN\t\t\tSAID\tContract\tStart\tFinish\tWarranty\tStart\tFinish')
    for row in input_data:
        print(row[0] + '\t' + row[1] + '\t' + row[2] + '\t' + row[3] + '\t' + row[4] + '\t'+ row[5] + '\t' + row[6] + '\t'+ row[7] + '\t' + row[8])


print_composed(composed)
outputpath = path.replace(path.split('\\')[-1], 'result.xlsx')
save(outputpath, composed)
browser.quit()
